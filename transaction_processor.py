import os
from datetime import datetime
import zipfile
from glob import glob
from openpyxl.reader.excel import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
import csv


# Change current working directory
PROJECT_FOLDER = r"C:\Users\Jaypr\iCloudDrive\coding_Lessons_And_Examples\Automate_The_Boring_Stuff\BudgetSheetProject"

# Destination file = the budget spreadsheet
DESTINATION_FILE = "ANNUAL-BUDGET 2026 (MAR - Present).xlsx"

# Add more rules here whenever you want
CATEGORY_RULES = [
    ("ZAXBY'S", "Eating out"),
    ("MCDONALD'S", "Eating out"),
    ("Bojangles", "Eating out"),
    ("JIMMY JOHNS", "Eating out"),
    ("KFC J718296", "Eating out"),
    ("To 52-Week ", "52-week Plan"),
    ("Erica Hamlin", "Sent to Eri adjust category"),
    ("UBER", "Ride Share"),
    ("UBR* PENDING.UBER.COM", "Ride Share"),
    ("LYFT   ", "Ride Share"),
    ("Sky Tobacco & Vape 3", "vapes"),
    ("PAPAS PACKAGE", "Alcohol"),
    ("COLUMBUS TOBACCO & VAP", "vapes"),
    ("WAL-MART ", "Groceries/household items"),
    ("PUBLIX ", "Groceries/household items"),
    ("WM SUPERCENTER", "Groceries/household items"),
    ("QT 1105 INSIDE", "Vending Machine/Gas station snack"),
    ("COCA COLA COLUMBUS GA", "Vending Machine/Gas station snack"),
    ("BEN HARM CH MINI MALL", "Vending Machine/Gas station snack"),
    ("TRACTOR SUPPLY", "Pets"),
    ("FUR-BABY PET SERVICES", "Pets"),
    ("LS MAXWELLS MOTORCYCLE", "Motorcycle Maintainece"),
    ("ATLANTA HARLEY-DAVIDSO", "Motorcycle Maintainece"),
    ("VS *WOW!", "Wow internet"),
    ("QT 720 OUTSIDE", "Gas"),
    ("SHELL SERVICE", "Gas"),
    ("SHELL", "Gas"),
    ("RACEWAY", "Gas"),
    ("MARATHON", "Gas"),
    ("BP", "Gas"),
    ("CIRCLE K", "Gas"),
    ("Prime Video Channels", "Prime Video"),
    ("AMAZON PRIME PMTS", "Amazon Prime Yearly"),
    ("Google One", "Google one"),
    ("DFAS-IN  IND, IN", "Monthly Pay"),
    ("Interest earned", "Refunds/PayBack"),
    ("Adobe", "Adobe"),
    ("CASH APP*DEEZY 2 FADED", "Haircut"),
    ("Mobile Check Deposit", "Refunds/PayBack"),

]

# Special transfer keywords
TRANSFER_KEYWORDS = [
    "From Savings/Emergency Vault",
    "From Early Pay Holding Cell Vault",
    "From Savings - 3475",
    "From weekly spending Vault",
    "From Savings - 7658"
]

TO_TRANSFER_KEYWORDS = [
    "To Checking - 5652",
    "To Early Pay Holding Cell Vault",
]

def get_to_transfer_deposit(description_text: str) -> str:
    """
    Extracts and returns the meaningful part of a deposit description.

    This function processes a given deposit description text and removes
    any leading "To" prefix, if present. The resulting text is then
    stripped of unnecessary whitespace before being returned.

    Args:
        description_text (str): The deposit description text to process.

    Returns:
        str: The processed deposit description text after removing the "To"
             prefix, if applicable, and trimming whitespace.
    """
    if description_text.startswith("To "):
        return description_text[3:].strip()
    return description_text.strip()

def get_effective_cell(ws, row, column):
    """
    Return a writable cell. If the requested cell is inside a merged range,
    return the top-left anchor cell of that merged range.
    """
    cell = ws.cell(row=row, column=column)

    if not isinstance(cell, MergedCell):
        return cell

    target_coord = f"{get_column_letter(column)}{row}"
    for merged_range in ws.merged_cells.ranges:
        if target_coord in merged_range:
            min_col, min_row, _, _ = merged_range.bounds
            return ws.cell(row=min_row, column=min_col)

    return cell

def is_valid_excel_file(file_path: str) -> bool:
    if not os.path.isfile(file_path):
        return False
    if not zipfile.is_zipfile(file_path):
        return False
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        wb.close()
        return True
    except Exception:
        return False

def get_newest_valid_source_file(folder: str, destination_file: str) -> str:
    source_files = [
        file_name for file_name in glob(os.path.join(folder, "*.xlsx"))
        if os.path.basename(file_name) != destination_file and not os.path.basename(file_name).startswith("~$")
    ]

    valid_files = [file_name for file_name in source_files if is_valid_excel_file(file_name)]

    if not valid_files:
        raise FileNotFoundError(
            "No valid Excel source workbook found. Make sure the downloaded file is a real .xlsx workbook."
        )

    return max(valid_files, key=os.path.getmtime)

def normalize_date(value):
    """
    Convert date-like values into something sortable.
    Invalid or missing dates get pushed to the end.
    """
    if value is None:
        return datetime.max

    if isinstance(value, datetime):
        return value

    try:
        return datetime.combine(value, datetime.min.time())
    except Exception:
        pass

    if isinstance(value, str):
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m/%d/%Y %H:%M:%S"):
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue

    return datetime.max

def normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower()

def get_header_column(source_headers, *possible_names):
    for name in possible_names:
        key = name.strip().lower()
        if key in source_headers:
            return source_headers[key]
    raise KeyError(f"Missing required column. Expected one of: {', '.join(possible_names)}")

def read_csv_rows(source_file):
    with open(source_file, mode="r", newline="", encoding="utf-8-sig") as csv_file:
        reader = csv.DictReader(csv_file)
        rows = []

        for row in reader:
            rows.append(row)

    return rows

def detect_budget_name(description_text, amount_value):
    budget_name = "MISC"

    if description_text.startswith("APPLE.COM/BILL") and round(amount_value, 2) == 5.99:
        budget_name = "Apple Music"
    elif description_text.startswith("APPLE.COM/BILL") and round(amount_value, 2) == 2.99:
        budget_name = "iCloud +"
    else:
        for text_to_match, category_name in CATEGORY_RULES:
            if description_text.startswith(text_to_match):
                budget_name = category_name
                break

    return budget_name

# Main budget Destination columns
DATE_COL = 8          # H
BUDGET_NAME_COL = 9   # I
AMOUNT_COL = 11       # K
ACCOUNT_COL = 15      # O
DESCRIPTION_COL = 16  # P
START_ROW = 69

# Special transfer section columns
TRANSFER_START_ROW = 105
TRANSFER_DATE_COL = 2       # B
TRANSFER_WITHDRAW_COL = 3   # C
TRANSFER_DEPOSIT_COL = 4    # D
TRANSFER_AMOUNT_COL = 5     # E

# Available account choices
ACCOUNT_CHOICES = [
    "SoFi Checking (2695)",
    "SoFi Savings (3475)",
    "Early Pay holding cell",
    "SoFi Vault (Savings/Emergency)"
]

# Available month sheet choices
MONTH_CHOICES = [
    "JAN", "FEB", "MAR", "APR", "MAY", "JUN",
    "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"
]

def process_transactions(source_file, account_name, destination_sheet_name):
    is_csv = source_file.endswith(".csv")

    if not is_csv:
        source_wb = load_workbook(source_file)
        source_ws = source_wb.active

    rows_to_import = []

    if is_csv:
        csv_rows = read_csv_rows(source_file)

        for row in csv_rows:
            date_text = str(row.get("Date")).strip()

            try:
                date_value = datetime.strptime(date_text, "%Y-%m-%d")
            except ValueError:
                try:
                    date_value = datetime.strptime(date_text, "%m/%d/%Y")
                except ValueError:
                    continue

            description_value = row.get("Description")
            amount_value = row.get("Amount")

            if description_value is None or amount_value is None:
                continue

            description_text = str(description_value).strip()

            try:
                amount_value = abs(float(amount_value))
            except (TypeError, ValueError):
                continue

            transaction_key = (date_value, amount_value, description_text)

            #if transaction_key in existing_transactions:
             #   continue

            budget_name = "MISC"

            if description_text.startswith("APPLE.COM/BILL") and round(amount_value, 2) == 5.99:
                budget_name = "Apple Music"
            elif description_text.startswith("APPLE.COM/BILL") and round(amount_value, 2) == 2.99:
                budget_name = "iCloud +"
            else:
                for text_to_match, category_name in CATEGORY_RULES:
                    if description_text.startswith(text_to_match):
                        budget_name = category_name
                        break

            rows_to_import.append({
                "date": date_value,
                "budget_name": budget_name,
                "amount": amount_value,
                "description": description_text,
                "account": account_name
            })

    else:
        source_headers = {}

        for cell in source_ws[1]:
            if cell.value is not None:
                header_name = normalize_header(cell.value)
                source_headers[header_name] = cell.column

        date_col = get_header_column(source_headers, "date", "posted date", "transaction date")
        description_col = get_header_column(source_headers, "description", "transaction description", "memo")
        amount_col = get_header_column(source_headers, "amount", "debit", "credit")

        print("Date column:", date_col)
        print("Description column:", description_col)
        print("Amount column:", amount_col)

        for row in range(2, source_ws.max_row + 1):
            date_value = source_ws.cell(row=row, column=date_col).value
            description_value = source_ws.cell(row=row, column=description_col).value
            amount_value = source_ws.cell(row=row, column=amount_col).value

            if description_value is None or amount_value is None:
                continue

            description_text = str(description_value).strip()

            try:
                amount_value = abs(float(amount_value))
            except (TypeError, ValueError):
                continue

            transaction_key = (date_value, amount_value, description_text)

            #if transaction_key in existing_transactions:
            #    continue

            budget_name = "MISC"

            if description_text.startswith("APPLE.COM/BILL") and round(amount_value, 2) == 5.99:
                budget_name = "Apple Music"
            elif description_text.startswith("APPLE.COM/BILL") and round(amount_value, 2) == 2.99:
                budget_name = "iCloud +"
            else:
                for text_to_match, category_name in CATEGORY_RULES:
                    if description_text.startswith(text_to_match):
                        budget_name = category_name
                        break

            rows_to_import.append({
                "date": date_value,
                "budget_name": budget_name,
                "amount": amount_value,
                "description": description_text,
                "account": account_name
            })

    # =========================
    # SORT TRANSACTIONS
    # =========================
    rows_to_import.sort(key=lambda item: normalize_date(item["date"]))

    # =========================
    # FIND FIRST EMPTY ROW
    # =========================
    # main_destination_row = START_ROW
    #
    # while destination_ws.cell(row=main_destination_row, column=DATE_COL).value is not None:
    #     main_destination_row += 1
    #
    # print("Starting write at row:", main_destination_row)
    #
    # # =========================
    # # WRITE TRANSACTIONS
    # # =========================
    # for item in rows_to_import:
    #     destination_ws.cell(row=main_destination_row, column=DATE_COL).value = item["date"]
    #     destination_ws.cell(row=main_destination_row, column=BUDGET_NAME_COL).value = item["budget_name"]
    #     destination_ws.cell(row=main_destination_row, column=AMOUNT_COL).value = item["amount"]
    #     destination_ws.cell(row=main_destination_row, column=ACCOUNT_COL).value = account_name
    #     destination_ws.cell(row=main_destination_row, column=DESCRIPTION_COL).value = item["description"]
    #
    #     main_destination_row += 1
    #
    # destination_wb.save(destination_path)
    #
    # print("Saved workbook:", destination_path)
    # print("Rows collected:", len(rows_to_import))

    return rows_to_import



